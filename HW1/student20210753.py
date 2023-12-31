#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook(filename = 'student.xlsx')
ws = wb.active
total_list = []
sheet_ranges = wb['Sheet1']
for i in range(2, sheet_ranges.max_row + 1):
    total = 0
    for j in range(4):
        if(j == 0):
            total += (sheet_ranges['C' + str(i)].value * 0.3)
        elif(j == 1):
             total += (sheet_ranges['D' + str(i)].value * 0.35)
        elif(j == 2):
            total += (sheet_ranges['E' + str(i)].value * 0.34)
        elif(j == 3):
            total += (sheet_ranges['F' + str(i)].value * 0.01)
    
    sheet_ranges['G' + str(i)].value = total
    total_list.append(total)

total_list.sort(reverse=True)

total_count = len(total_list)

a_range = int(total_count * 0.3)
a_grade = total_list[:a_range]
aPlus_range = int(len(a_grade) * 0.5)
aPlus_grade = a_grade[:aPlus_range]
a_grade = a_grade[aPlus_range:]

b_range = int(total_count * 0.7)
b_grade = total_list[a_range:b_range]
bPlus_range = int(len(b_grade) * 0.5)
bPlus_grade = b_grade[:bPlus_range]
b_grade = b_grade[bPlus_range:]

c_grade = total_list[b_range:]
cPlus_range = int(len(c_grade) * 0.5)
cPlus_grade = c_grade[:cPlus_range]
c_grade = c_grade[cPlus_range:]

for i in range(2, sheet_ranges.max_row + 1):
    if sheet_ranges['G' + str(i)].value < 40:
        grade = 'F'
    elif sheet_ranges['G' + str(i)].value in aPlus_grade:
        grade = 'A+'
    elif sheet_ranges['G' + str(i)].value in a_grade:
        grade = 'A'
    elif sheet_ranges['G' + str(i)].value in bPlus_grade:
        grade = 'B+'
    elif sheet_ranges['G' + str(i)].value in b_grade:
        grade = 'B'
    elif sheet_ranges['G' + str(i)].value in cPlus_grade:
        grade = 'C+'
    elif sheet_ranges['G' + str(i)].value in c_grade:
        grade = 'C'

    sheet_ranges['H' + str(i)].value = grade

wb.save(filename = 'student.xlsx')
